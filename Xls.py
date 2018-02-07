# -*- coding: utf-8 -*-
# author lis
# http://www.lisz-works.com

import openpyxl as xl
import Util

class Xls:
    book = None
    sheet = None
    data = []
    bookName = ""
    # debug var ----------
    isTraceback = False#True

    def __init__(self):
        self.dbgPrintTraceback("__init__()")
        self.initialize()

    def __init__(self, bookName):
        self.dbgPrintTraceback("__init__(bookName)")
        self.initialize()
        self.openBook(bookName)

    def __init__(self, bookName, sheetName):
        self.dbgPrintTraceback("__init__(bookName, sheetName)")
        self.initialize()
        if self.openBook(bookName):
            self.openSheet(sheetName)

    def initialize(self):
        self.dbgPrintTraceback("initialize")

    def openBook(self, bookName, isDataOnly=True):
        self.dbgPrintTraceback("openBook")
        if not Util.chkString(bookName):
            return False
        self.book = xl.load_workbook(bookName, data_only=isDataOnly)
        #------ debug
        #print("*<type>book: " + str(type(self.book)))
        #------ debug
        if not self.isOpenedBook(self.book, True):
            return False
        self.bookName = bookName
        return True

    def openSheet(self, sheetName):
        self.dbgPrintTraceback("openSheet")
        if not Util.chkString(sheetName):
            return False
        if not self.isOpenedBook(self.book):
            return False
        self.sheet = self.book.get_sheet_by_name(sheetName)
        return self.isOpenedSheet(self.sheet, True)

    def createSheet(self, sheetName, renew=False):
        self.dbgPrintTraceback("createSheet")
        if not Util.chkString(sheetName):
            return False
        if renew:
            if self.existSheet(sheetName):
                # シートが既に存在する場合、削除
                s = self.book.get_sheet_by_name(sheetName)
                self.book.remove_sheet(s)
        self.sheet = self.book.create_sheet(title=sheetName)
        return self.isOpenedSheet(self.sheet, True)

    def resultText(self, result, text):
        self.dbgPrintTraceback("resultText")
        log = "[SUCCESS]" if ( result ) else "[FAILED]"
        print(log + text)

    # opened=True
    def isOpenedBook(self, book, isOutputLog=False):
        self.dbgPrintTraceback("isOpenedBook")
        ret = not book is None
        #------ debug
        #print("<type>book: " + str(type(book)) + ", not book is None = " + str(ret))
        #------ debug
        if isOutputLog:
            self.resultText(ret, "open book")
        return ret

    # opened=True
    def isOpenedSheet(self, sheet, isOutputLog=False):
        self.dbgPrintTraceback("isOpenedSheet")
        ret = not sheet is None
        if isOutputLog:
            self.resultText(ret, "open sheet")
        return ret

    # opened=True
    def isOpened(self, book, sheet, isOutputLog=False):
        self.dbgPrintTraceback("isOpened")
        if not self.isOpenedBook(book, isOutputLog):
            return False
        if not self.isOpenedSheet(sheet, isOutputLog):
            return False
        #------ debug
        #print("<type>book: " + str(type(book)))
        #print("<type>sheet: " + str(type(sheet)))
        #------ debug
        return True

    # 対象のシートが存在する=True
    def existSheet(self, sheetName):
        sheets = self.book.get_sheet_names()
        return sheetName in sheets

    def getCellValue(self, _row=0, _col=0):
        self.dbgPrintTraceback("getCellValue")
        if _row <= 0 or _col <= 0:
            return ""
        #------ debug
        #print("isOpened = " + str(self.isOpened(self.book, self.sheet)))
        #print("r,c = " + str(_row) + ", " + str(_col))
        #print("self.book is None = " + str(self.book is None))
        #print("self.sheet is None = " + str(self.sheet is None))
        #print("- - - - - - - - - - -")
        #------ debug
        return self.sheet.cell(row=_row, column=_col).value

    # 空=True
    def isBlankCell(self, val):
        self.dbgPrintTraceback("isBlankCell: val=" + str(val))
        if val is None:
            return True
        return (len(str(val)) <= 0)

    # 開始セルから行列共に空白セルとなるまでデータを全てロードする
    # 先頭セルは、idx=1
    def loadAllData(self, startRow=0, startCol=0):
        self.dbgPrintTraceback("loadAllData")
        if not self.isOpened(self.book, self.sheet, False):
            return False
        row = startRow
        col = startCol
        val = self.getCellValue(row, col)
        log = ""
        while not self.isBlankCell(val):
            log = ""
            list = []
            while not self.isBlankCell(val):
                # 1行分の値をlist[]にセットする
                log += str(val) + ", " # is debug
                list.append(val)
                col += 1
                val = self.getCellValue(row, col)
            # 1行分の値list[]をdata[]にセットする
            self.data.append(list)
            col = startCol
            row += 1
            val = self.getCellValue(row, col)
            #print(log)  # is debug
        print("[SUCCESS]All data reading")
        print(self.data)
        return True

    # 引数の配列を指定セルから横方向に展開
    def writeHorizontal(self, data, row=1, col=1):
        self.dbgPrintTraceback("writeHorizontal")
        if not Util.chkList(data):
            return False
        # 渡されたデータをRCから空になるまで書込
        c = col
        for val in data:
            self.sheet.cell(row=row, column=c).value = val
            c += 1
        print("[SUCCESS]write horizontal")
        return True

    # 保存
    def save(self):
        self.dbgPrintTraceback("save")
        self.book.save(self.bookName)
        print("[FILE SAVING]")

    # is debug ------------------------------------------------------
    def dbgPrintTraceback(self, log):
        if self.isTraceback:
            print("<function> " + log)

    def dbgPrint(self, log):
        print(log)

if __name__ == '__main__':
    x = Xls("data.xlsx", "data")
    x.loadAllData(1, 1)
